import subprocess, sys
subprocess.check_call([sys.executable, "-m", "pip", "install", "yfinance", "openpyxl", "pandas", "-q"])

import yfinance as yf
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
TICKERS = ["NWPX", "OMEX", "NVEC", "JRI", "FMBRX"]
START = "2020-12-01"
END   = "2026-01-01"

# ── Sheet 1: Monthly Close Price ─────────────────────────────────────────────
price_frames = {}
for ticker in TICKERS:
    try:
        df = yf.download(ticker, start=START, end=END, auto_adjust=False, progress=False)
        if df.empty:
            print(f"[Price] Warning: No data for {ticker}")
            price_frames[ticker] = pd.Series(dtype=float)
            continue
        close = df["Close"]
        if isinstance(close, pd.DataFrame):
            close = close.iloc[:, 0]
        monthly = close.resample("ME").last()
        monthly.index = monthly.index.to_period("M").to_timestamp("M")
        price_frames[ticker] = monthly
        print(f"✅ [Price] {ticker}: last = {monthly.iloc[-1]:.4f} ({monthly.index[-1].strftime('%b %d %Y')})")
    except Exception as e:
        print(f"[Price] Error {ticker}: {e}")
        price_frames[ticker] = pd.Series(dtype=float)

all_dates = pd.date_range("2020-12-31", "2025-12-31", freq="ME")
price_combined = pd.DataFrame(index=all_dates)
price_combined.index.name = "Date"
for ticker, series in price_frames.items():
    price_combined[ticker] = series
price_combined = price_combined.loc["2020-12-01":"2025-12-31"]

# ── Sheet 2: Monthly Dividends ────────────────────────────────────────────────
div_frames = {}
for ticker in TICKERS:
    try:
        tk = yf.Ticker(ticker)
        divs = tk.dividends  # DatetimeIndex with tz
        if divs.empty:
            print(f"[Div] No dividends for {ticker}")
            div_frames[ticker] = pd.Series(dtype=float)
            continue
        # Remove timezone, filter period
        divs.index = divs.index.tz_localize(None)
        divs = divs.loc[START:END]
        # Sum dividends per month (in case multiple in same month)
        monthly_div = divs.resample("ME").sum()
        monthly_div = monthly_div[monthly_div > 0]
        monthly_div.index = monthly_div.index.to_period("M").to_timestamp("M")
        div_frames[ticker] = monthly_div
        print(f"✅ [Div] {ticker}: {len(monthly_div)} months with dividends")
    except Exception as e:
        print(f"[Div] Error {ticker}: {e}")
        div_frames[ticker] = pd.Series(dtype=float)

div_combined = pd.DataFrame(index=all_dates)
div_combined.index.name = "Date"
for ticker, series in div_frames.items():
    div_combined[ticker] = series
div_combined = div_combined.loc["2020-12-01":"2025-12-31"]
# Replace 0 with NaN so empty months show as blank
div_combined = div_combined.replace(0, pd.NA)

# ── Build Excel ───────────────────────────────────────────────────────────────
wb = Workbook()

def style_sheet(ws, combined, title_note):
    HEADER_FILL  = PatternFill("solid", start_color="1F4E79")
    HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    DATE_FILL    = PatternFill("solid", start_color="D6E4F0")
    DATE_FONT    = Font(name="Arial", bold=True, size=10)
    DATA_FONT    = Font(name="Arial", size=10)
    ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
    ALIGN_RIGHT  = Alignment(horizontal="right", vertical="center")
    thin   = Side(style="thin", color="BDD7EE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["Date"] + TICKERS
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = ALIGN_CENTER
        cell.border = border

    for row_idx, (date, row) in enumerate(combined.iterrows(), start=2):
        date_cell = ws.cell(row=row_idx, column=1, value=date.strftime("%b-%Y"))
        date_cell.font = DATE_FONT
        date_cell.fill = DATE_FILL
        date_cell.alignment = ALIGN_CENTER
        date_cell.border = border

        for col_idx, ticker in enumerate(TICKERS, start=2):
            val  = row[ticker]
            cell = ws.cell(row=row_idx, column=col_idx)
            if pd.notna(val) and val != 0:
                cell.value = round(float(val), 4)
                cell.number_format = '#,##0.0000'
            else:
                cell.value = None
            cell.font = DATA_FONT
            cell.alignment = ALIGN_RIGHT
            cell.border = border
            if row_idx % 2 == 0:
                cell.fill = PatternFill("solid", start_color="EBF3FB")

    ws.column_dimensions["A"].width = 14
    for col_idx in range(2, len(TICKERS) + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14
    ws.freeze_panes = "A2"

    note_row = len(combined) + 3
    note_cell = ws.cell(row=note_row, column=1, value=title_note)
    note_cell.font = Font(name="Arial", italic=True, color="808080", size=9)
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=len(TICKERS)+1)

# Sheet 1 - Price
ws1 = wb.active
ws1.title = "Monthly Close Price"
style_sheet(ws1, price_combined,
    f"Source: Yahoo Finance | Raw Close Price (non-adjusted) | Dec 2020 - Dec 2025 | Retrieved: {pd.Timestamp.today().strftime('%Y-%m-%d')}")

# Sheet 2 - Dividends
ws2 = wb.create_sheet("Monthly Dividends")
style_sheet(ws2, div_combined,
    f"Source: Yahoo Finance | Dividends summed per month | Dec 2020 - Dec 2025 | Retrieved: {pd.Timestamp.today().strftime('%Y-%m-%d')}")

output_path = os.path.join(SCRIPT_DIR, "Monthly_Close_Price.xlsx")
wb.save(output_path)
print(f"\n✅ 完成！Excel 文件已保存到：")
print(f"   {output_path}")
print(f"   Sheet 1: Monthly Close Price")
print(f"   Sheet 2: Monthly Dividends")
